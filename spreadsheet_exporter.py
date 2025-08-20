import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from config import SPREADSHEET_PATH

class SpreadsheetExporter:
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Greensboro Events"
    
    def export_events(self, events_df):
        """Export events DataFrame to Excel with formatting"""
        if events_df.empty:
            print("No events to export")
            return
        
        # Add title and metadata
        self._add_header()
        
        # Add data starting from row 4
        self._add_data(events_df)
        
        # Format the spreadsheet
        self._format_spreadsheet()
        
        # Save the file
        self.workbook.save(SPREADSHEET_PATH)
        print(f"Events exported to {SPREADSHEET_PATH}")
    
    def _add_header(self):
        """Add header information to the spreadsheet"""
        # Title
        self.worksheet['A1'] = "Greensboro, NC Events"
        self.worksheet['A1'].font = Font(size=16, bold=True)
        
        # Generation timestamp
        self.worksheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        self.worksheet['A2'].font = Font(size=10, italic=True)
    
    def _add_data(self, events_df):
        """Add event data to the spreadsheet"""
        # Add column headers
        headers = list(events_df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=4, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, (_, row_data) in enumerate(events_df.iterrows(), 5):
            for col_num, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num, value=value)
                
                # Special formatting for URLs
                if col_num in [7, 8]:  # Image URL and Event URL columns
                    if value and str(value).startswith('http'):
                        cell.hyperlink = value
                        cell.font = Font(color="0000FF", underline="single")
    
    def _format_spreadsheet(self):
        """Apply formatting to the spreadsheet"""
        # Auto-adjust column widths
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders and alternating row colors
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to data area
        max_row = self.worksheet.max_row
        max_col = self.worksheet.max_column
        
        for row in range(4, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    def create_summary_sheet(self, events_df):
        """Create a summary sheet with statistics"""
        summary_sheet = self.workbook.create_sheet("Summary")
        
        # Event count by type
        if 'Event Type' in events_df.columns:
            event_type_counts = events_df['Event Type'].value_counts()
            
            summary_sheet['A1'] = "Event Summary"
            summary_sheet['A1'].font = Font(size=14, bold=True)
            
            summary_sheet['A3'] = "Events by Type:"
            summary_sheet['A3'].font = Font(bold=True)
            
            row = 4
            for event_type, count in event_type_counts.items():
                summary_sheet[f'A{row}'] = event_type
                summary_sheet[f'B{row}'] = count
                row += 1
        
        # Total events
        summary_sheet['A2'] = f"Total Events: {len(events_df)}"

if __name__ == "__main__":
    # Test export
    sample_data = pd.DataFrame({
        'Title': ['Test Event'],
        'Date': ['2023-12-25'],
        'Location': ['Greensboro, NC']
    })
    
    exporter = SpreadsheetExporter()
    exporter.export_events(sample_data)
